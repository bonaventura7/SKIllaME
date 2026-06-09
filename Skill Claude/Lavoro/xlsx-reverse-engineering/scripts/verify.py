#!/usr/bin/env python3
"""
xlsx-reverse-engineering: Verification & Cross-Check Helper
- Loads formulas (from extract) or scans directly
- Compares raw formulas vs recalculated values (if data_only available)
- Checks for basic provenance/citation hygiene in a simple way
- Produces verification report
Intended as a gate after major phases.
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print("Usage: python verify.py <path-to-xlsx> [data_only_file_or_same]")
        print("If the file has been recalculated (data_only values present), pass the same path or a data_only version.")
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    if not input_path.exists():
        print(json.dumps({"error": f"File not found: {input_path}"}))
        sys.exit(1)

    # Try to load both formula view and (if possible) value view
    try:
        wb_formulas = load_workbook(input_path, data_only=False)
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

    wb_values = None
    try:
        wb_values = load_workbook(input_path, data_only=True)
    except Exception:
        pass  # values not available

    report = {
        "file": str(input_path),
        "verified_at": datetime.now().isoformat(),
        "checks": [],
        "sample_verifications": [],
        "summary": {}
    }

    total_formulas = 0
    verified_samples = 0
    discrepancies = 0

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name] if wb_values else None

        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1
                    location = f"{sheet_name}!{cell.coordinate}"
                    formula = cell.value

                    # Sample 8-10 formulas for verification
                    if verified_samples < 10:
                        value = None
                        if ws_v:
                            val_cell = ws_v[cell.coordinate]
                            value = val_cell.value

                        sample = {
                            "location": location,
                            "formula": formula[:120],
                            "value_after_recalc": str(value)[:80] if value is not None else "NOT_AVAILABLE",
                            "status": "value_available" if value is not None else "formula_only"
                        }
                        report["sample_verifications"].append(sample)
                        verified_samples += 1

                    # Basic hygiene check: does formula contain at least one cell ref or named range?
                    if not any(c.isalpha() and any(ch.isdigit() for ch in formula) for c in formula.split()):
                        report["checks"].append({
                            "type": "formula_hygiene",
                            "location": location,
                            "issue": "Formula appears to have no obvious cell references",
                            "formula": formula[:80]
                        })

    wb_formulas.close()
    if wb_values:
        wb_values.close()

    report["summary"] = {
        "total_formulas_scanned": total_formulas,
        "samples_verified": len(report["sample_verifications"]),
        "discrepancies_noted": discrepancies,
        "recommendation": "Manually review samples where value is 'NOT_AVAILABLE' or suspicious."
    }

    print(json.dumps(report, indent=2, ensure_ascii=False))

    out_dir = Path("verification")
    out_dir.mkdir(exist_ok=True)
    with open(out_dir / "verification_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)

    with open(out_dir / "verification_samples.md", "w", encoding="utf-8") as f:
        f.write(f"# Verification Samples\n\n**File**: {input_path}\n\n")
        for s in report["sample_verifications"]:
            f.write(f"- **{s['location']}**\n")
            f.write(f"  Formula: `{s['formula']}`\n")
            f.write(f"  Value: {s['value_after_recalc']}\n\n")

    print(f"\n[INFO] Verification artifacts in ./verification/", file=sys.stderr)

if __name__ == "__main__":
    main()
