#!/usr/bin/env python3
"""
formula_decomposer.py — Recursive Formula Decomposition Engine
Breaks down complex nested formulas into hierarchical steps for opacity reduction and business logic mapping.
Based on engines/formula-tracer.md and formula-patterns.md
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def decompose_formula(formula: str, depth: int = 0) -> dict:
    """Recursively decompose a formula into structure."""
    formula = formula.strip()
    if not formula.startswith('='):
        return {"type": "literal", "value": formula, "depth": depth}

    # Remove leading =
    inner = formula[1:]

    # Detect top-level function
    match = re.match(r'^([A-Z_]+)\((.*)\)$', inner, re.IGNORECASE)
    if match:
        func = match.group(1).upper()
        args_str = match.group(2)
        args = split_top_level_args(args_str)
        decomposed_args = [decompose_formula(arg.strip(), depth + 1) for arg in args]
        return {
            "type": "function",
            "function": func,
            "depth": depth,
            "args": decomposed_args,
            "original": formula
        }
    else:
        # Fallback for expressions or array formulas
        return {
            "type": "expression",
            "value": inner,
            "depth": depth,
            "original": formula
        }

def split_top_level_args(args_str: str) -> list:
    """Split arguments respecting nested parentheses and quotes."""
    args = []
    current = ""
    depth = 0
    in_quote = False
    for char in args_str:
        if char == '"' and not in_quote:
            in_quote = True
        elif char == '"' and in_quote:
            in_quote = False
        elif char == '(' and not in_quote:
            depth += 1
        elif char == ')' and not in_quote:
            depth -= 1
        elif char == ',' and depth == 0 and not in_quote:
            args.append(current.strip())
            current = ""
            continue
        current += char
    if current.strip():
        args.append(current.strip())
    return args

def analyze_formula(formula: str) -> dict:
    """Full analysis: decompose + classify + risk."""
    tree = decompose_formula(formula)
    complexity = len(re.findall(r'\(', formula))
    volatile = any(v in formula.upper() for v in ['INDIRECT', 'OFFSET', 'NOW', 'RAND'])
    risk = "Critical" if complexity > 8 or volatile else ("High" if complexity > 4 else "Medium")
    return {
        "original": formula,
        "decomposition_tree": tree,
        "complexity_score": complexity,
        "risk_level": risk,
        "is_volatile": volatile
    }

def main():
    if len(sys.argv) < 2:
        print("Usage: python formula_decomposer.py <file.xlsx> [--sheet Sheet1] [--cell A1] [--pretty]")
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    sheet = None
    cell = None
    pretty = "--pretty" in sys.argv

    for i, arg in enumerate(sys.argv):
        if arg == "--sheet" and i+1 < len(sys.argv):
            sheet = sys.argv[i+1]
        if arg == "--cell" and i+1 < len(sys.argv):
            cell = sys.argv[i+1]

    wb = load_workbook(input_path, data_only=False)
    results = []

    if sheet and cell:
        ws = wb[sheet]
        f = ws[cell].value
        if isinstance(f, str) and f.startswith('='):
            results.append(analyze_formula(f))
    else:
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith('='):
                        res = analyze_formula(c.value)
                        res["location"] = f"{sname}!{c.coordinate}"
                        results.append(res)
                        if len(results) > 50:  # cap for demo
                            break

    output = {
        "file": str(input_path),
        "decomposed_at": datetime.now().isoformat(),
        "formulas_analyzed": len(results),
        "results": results[:30]  # limit output
    }

    if pretty:
        print(json.dumps(output, indent=2, ensure_ascii=False))
    else:
        print(json.dumps(output))

    wb.close()

if __name__ == "__main__":
    main()
