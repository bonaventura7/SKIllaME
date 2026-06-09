#!/usr/bin/env python3
"""
formula_audit.py — Deep Formula Analysis (Phase 2)
Supports --data-only for value comparison.
"""

import json
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

VOLATILE = ["INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN"]
MODERN365 = ["XLOOKUP", "FILTER", "UNIQUE", "SORT", "TRANSPOSE", "LAMBDA", "LET"]

def complexity_score(formula: str) -> int:
    score = 0
    score += formula.count("(") * 2
    score += len(re.findall(r'\bIF\b', formula, re.I)) * 3
    score += len(re.findall(r'\bSUMPRODUCT\b', formula, re.I)) * 5
    score += len(re.findall(r'\bINDEX\b|\bMATCH\b', formula, re.I)) * 4
    score += len(re.findall(r'\+|\-|\*|\/', formula))
    return min(100, max(1, score))

def extract_refs(formula: str, current_sheet: str):
    refs = []
    for match in re.finditer(r"'?([^'!]+)'?!?([A-Z]{1,3}\$?\d+)", formula):
        refs.append({"sheet": match.group(1), "cell": match.group(2)})
    for match in re.finditer(r'\b([A-Z]{1,3}\$?\d+)\b', formula):
        refs.append({"sheet": current_sheet, "cell": match.group(1)})
    return refs

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    parser.add_argument("--data-only", action="store_true")
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()

    input_path = Path(args.file).expanduser().resolve()
    wb = load_workbook(input_path, data_only=False)
    wb_values = load_workbook(input_path, data_only=True) if args.data_only else None

    result = {
        "file": str(input_path),
        "data_only_mode": args.data_only,
        "scanned_at": datetime.now().isoformat(),
        "formula_summary": {"total": 0, "avg_complexity": 0, "high_complexity_count": 0},
        "complex_formulas": [],
        "issues": [],
        "volatile_functions_detected": [],
        "modern365_functions_detected": [],
        "dependency_graph": {},
        "function_usage": defaultdict(int)
    }

    all_complexities = []
    dep_graph = defaultdict(list)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_v = wb_values[sheet_name] if wb_values else None

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    location = f"{sheet_name}!{cell.coordinate}"
                    result["formula_summary"]["total"] += 1

                    score = complexity_score(formula)
                    all_complexities.append(score)

                    if score > 15:
                        val = None
                        if ws_v:
                            try:
                                val = ws_v[cell.coordinate].value
                            except:
                                pass
                        result["complex_formulas"].append({
                            "location": location,
                            "complexity": score,
                            "formula": formula[:140],
                            "value": str(val)[:40] if val is not None else None
                        })

                    upper = formula.upper()
                    for v in VOLATILE:
                        if v in upper and v not in result["volatile_functions_detected"]:
                            result["volatile_functions_detected"].append(v)
                    for m in MODERN365:
                        if m in upper and m not in result["modern365_functions_detected"]:
                            result["modern365_functions_detected"].append(m)

                    for func in re.findall(r'\b([A-Z_]+)\(', upper):
                        result["function_usage"][func] += 1

                    refs = extract_refs(formula, sheet_name)
                    for ref in refs:
                        target = f"{ref['sheet']}!{ref['cell']}"
                        dep_graph[target].append(location)

    result["dependency_graph"] = {k: v[:8] for k, v in list(dep_graph.items())[:150]}

    if all_complexities:
        result["formula_summary"]["avg_complexity"] = round(sum(all_complexities) / len(all_complexities), 1)
        result["formula_summary"]["high_complexity_count"] = len(result["complex_formulas"])

    result["function_usage"] = dict(sorted(result["function_usage"].items(), key=lambda x: -x[1])[:12])

    wb.close()
    if wb_values:
        wb_values.close()

    if args.pretty:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(json.dumps(result))

if __name__ == "__main__":
    main()
