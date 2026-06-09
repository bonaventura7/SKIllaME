#!/usr/bin/env python3
"""
xlsx-reverse-engineering: Phase 0/1 Inventory Script
Produces structured inventory of sheets, visibility, protections, defined names, and high-level stats.
Designed to be run early and often. Outputs JSON + writes helpful side files.
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) < 2:
        print("Usage: python inventory.py <path-to-xlsx-or-xlsm>")
        print("Outputs JSON to stdout and creates ./inventory/ artifacts next to the script or cwd.")
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    if not input_path.exists():
        print(json.dumps({"error": f"File not found: {input_path}"}))
        sys.exit(1)

    try:
        wb = load_workbook(input_path, data_only=False, read_only=False)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open with openpyxl: {str(e)}"}))
        sys.exit(1)

    inventory = {
        "file": {
            "path": str(input_path),
            "name": input_path.name,
            "size_bytes": input_path.stat().st_size,
            "extension": input_path.suffix.lower(),
            "scanned_at": datetime.now().isoformat()
        },
        "workbook": {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "active_sheet": wb.active.title if wb.active else None,
            "defined_names_count": len(wb.defined_names),
            "has_vba": input_path.suffix.lower() == ".xlsm"
        },
        "sheets": [],
        "defined_names": [],
        "summary": {}
    }

    # Per-sheet details
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Visibility (very important for RE)
        state = getattr(ws, 'sheet_state', 'visible')
        if state not in ('visible', 'hidden', 'veryHidden'):
            state = 'visible'  # default

        # Dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        dimensions = f"{get_column_letter(1)}1:{get_column_letter(max_col)}{max_row}" if max_row and max_col else "empty"

        # Rough formula count (scan cells)
        formula_count = 0
        sample_formulas = []
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 500), max_col=min(max_col, 50)):  # bounded for speed
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    if len(sample_formulas) < 3:
                        sample_formulas.append({
                            "cell": cell.coordinate,
                            "formula": cell.value[:120] + ("..." if len(cell.value) > 120 else "")
                        })

        # Protection
        protection = {
            "sheet_protected": bool(getattr(ws, 'protection', None) and ws.protection.sheet),
            "password_set": bool(getattr(ws, 'protection', None) and ws.protection.password),
        }

        sheet_info = {
            "name": sheet_name,
            "state": state,  # visible | hidden | veryHidden
            "dimensions": dimensions,
            "max_row": max_row,
            "max_col": max_col,
            "formula_count_sampled": formula_count,
            "sample_formulas": sample_formulas,
            "protection": protection,
            "has_tables": len(getattr(ws, 'tables', {})) > 0,
        }
        inventory["sheets"].append(sheet_info)

    # Defined names (named ranges / variables) - FIXED for current openpyxl
    for name_str in wb.defined_names:
        dn_obj = wb.defined_names[name_str]
        dn = {
            "name": getattr(dn_obj, "name", name_str),
            "attr_text": getattr(dn_obj, "attr_text", None),
            "hidden": bool(getattr(dn_obj, "hidden", False)),
            "value": str(getattr(dn_obj, "value", ""))[:200] if getattr(dn_obj, "value", None) else None,
        }
        inventory["defined_names"].append(dn)

    # Summary stats
    visible_count = sum(1 for s in inventory["sheets"] if s["state"] == "visible")
    hidden_count = sum(1 for s in inventory["sheets"] if s["state"] == "hidden")
    very_hidden_count = sum(1 for s in inventory["sheets"] if s["state"] == "veryHidden")
    total_formulas_sampled = sum(s["formula_count_sampled"] for s in inventory["sheets"])

    inventory["summary"] = {
        "visible_sheets": visible_count,
        "hidden_sheets": hidden_count,
        "very_hidden_sheets": very_hidden_count,
        "total_defined_names": len(inventory["defined_names"]),
        "hidden_named_ranges": sum(1 for n in inventory["defined_names"] if n["hidden"]),
        "total_formulas_sampled_across_sheets": total_formulas_sampled,
        "has_macros": inventory["workbook"]["has_vba"],
        "notes": "Formula counts are sampled (capped per sheet for performance). Run full extraction for complete list."
    }

    wb.close()

    # Output JSON to stdout (for agent to parse)
    print(json.dumps(inventory, indent=2, ensure_ascii=False))

    # Also write helpful files (idempotent, good for REVERSE-EXCEL.md consumption)
    out_dir = Path("inventory")
    out_dir.mkdir(exist_ok=True)

    # sheets.csv friendly
    with open(out_dir / "sheets.csv", "w", encoding="utf-8") as f:
        f.write("name,state,dimensions,max_row,max_col,formulas_sampled,protected\n")
        for s in inventory["sheets"]:
            f.write(f'"{s["name"]}","{s["state"]}","{s["dimensions"]}",{s["max_row"]},{s["max_col"]},{s["formula_count_sampled"]},{s["protection"]["sheet_protected"]}\n')

    # names.csv
    with open(out_dir / "names.csv", "w", encoding="utf-8") as f:
        f.write("name,hidden,value_preview\n")
        for n in inventory["defined_names"]:
            val = (n["value"] or "")[:80].replace('"', '""')
            f.write(f'"{n["name"]}",{n["hidden"]},"{val}"\n')

    # full json for later use
    with open(out_dir / "full_inventory.json", "w", encoding="utf-8") as f:
        json.dump(inventory, f, indent=2, ensure_ascii=False)

    print(f"\n[INFO] Side artifacts written to ./inventory/ (sheets.csv, names.csv, full_inventory.json)", file=sys.stderr)

if __name__ == "__main__":
    main()
