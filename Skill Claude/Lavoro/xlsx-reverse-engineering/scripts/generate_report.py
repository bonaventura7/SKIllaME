#!/usr/bin/env python3
"""
generate_report.py — Professional Documentation Generation (Phase 4)
Enhanced version:
- More sheets
- Professional formatting inspired by xlsx skill (headers, colors, number formats)
- Severity color coding
- Simple charts (bar for complexity, if possible)
- Integration note for recalc
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Professional styling (aligned with xlsx skill)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=12, color="2F5496")

SEVERITY_COLORS = {
    "high": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "medium": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "low": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row_num, start_col=1, end_col=6):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def create_report(input_file: str, output_file: str, inspect=None, audit=None, logic=None):
    wb = Workbook()

    # ========== Cover ==========
    ws = wb.active
    ws.title = "Cover"
    ws['A1'] = "EXCEL REVERSE ENGINEERING REPORT"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    ws['A3'] = "Source File:"
    ws['B3'] = input_file
    ws['A4'] = "Generated On:"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws['A5'] = "Analysis Tool:"
    ws['B5'] = "xlsx-reverse-engineering skill (4-phase pipeline)"

    ws['A7'] = "EXECUTIVE SUMMARY"
    ws['A7'].font = SUBTITLE_FONT

    if inspect:
        ws['A9'] = "Total Sheets:"
        ws['B9'] = inspect.get("summary", {}).get("total_sheets", "N/A")
        ws['A10'] = "Total Formulas:"
        ws['B10'] = inspect.get("summary", {}).get("total_formulas", "N/A")
        ws['A11'] = "Hidden Sheets:"
        ws['B11'] = inspect.get("summary", {}).get("hidden_sheets", 0) + inspect.get("summary", {}).get("very_hidden_sheets", 0)

    if audit:
        ws['A13'] = "Avg Formula Complexity:"
        ws['B13'] = audit.get("formula_summary", {}).get("avg_complexity", "N/A")
        ws['A14'] = "High Complexity Formulas:"
        ws['B14'] = audit.get("formula_summary", {}).get("high_complexity_count", 0)

    ws['A17'] = "Note: For verified calculated values, run the xlsx skill's recalc.py before analysis."
    ws['A17'].font = Font(italic=True, color="666666")

    # ========== Sheet Inventory ==========
    ws2 = wb.create_sheet("Sheet Inventory")
    ws2['A1'] = "Sheet Inventory"
    ws2['A1'].font = TITLE_FONT

    headers = ["Sheet Name", "Visibility", "Formula Count", "Has Tables", "Protected"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header_row(ws2, 3, 1, 5)

    if inspect and "sheets" in inspect:
        for i, s in enumerate(inspect["sheets"], 4):
            ws2.cell(row=i, column=1, value=s.get("name"))
            ws2.cell(row=i, column=2, value=s.get("state", "visible"))
            ws2.cell(row=i, column=3, value=s.get("formula_count", 0))
            ws2.cell(row=i, column=4, value=s.get("has_tables", False))
            ws2.cell(row=i, column=5, value=s.get("protected", False))

    # ========== Formula Analysis ==========
    ws3 = wb.create_sheet("Formula Analysis")
    ws3['A1'] = "Formula Analysis & Complexity"
    ws3['A1'].font = TITLE_FONT

    if audit:
        ws3['A3'] = "Total Formulas Analyzed"
        ws3['B3'] = audit.get("formula_summary", {}).get("total", 0)
        ws3['A4'] = "Average Complexity (1-100)"
        ws3['B4'] = audit.get("formula_summary", {}).get("avg_complexity", 0)
        ws3['A5'] = "High Complexity (>15)"
        ws3['B5'] = audit.get("formula_summary", {}).get("high_complexity_count", 0)

    # Top complex formulas
    ws3['A8'] = "Top Complex Formulas"
    style_header_row(ws3, 8, 1, 3)
    ws3.cell(row=8, column=1, value="Location")
    ws3.cell(row=8, column=2, value="Complexity")
    ws3.cell(row=8, column=3, value="Formula Preview")

    if audit and audit.get("complex_formulas"):
        for i, f in enumerate(audit["complex_formulas"][:15], 9):
            ws3.cell(row=i, column=1, value=f["location"])
            ws3.cell(row=i, column=2, value=f["complexity"])
            ws3.cell(row=i, column=3, value=f["formula"][:100])

    # ========== Issues & Warnings ==========
    ws4 = wb.create_sheet("Issues & Warnings")
    ws4['A1'] = "Issues & Warnings"
    ws4['A1'].font = TITLE_FONT

    headers4 = ["Severity", "Location", "Issue", "Recommendation"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=col, value=h)
    style_header_row(ws4, 3, 1, 4)

    if audit and audit.get("issues"):
        for i, issue in enumerate(audit["issues"], 4):
            sev = issue.get("severity", "medium")
            cell = ws4.cell(row=i, column=1, value=sev.upper())
            if sev in SEVERITY_COLORS:
                cell.fill = SEVERITY_COLORS[sev]
            ws4.cell(row=i, column=2, value=issue.get("location"))
            ws4.cell(row=i, column=3, value=issue.get("issue"))
            ws4.cell(row=i, column=4, value=issue.get("recommendation", ""))

    # ========== Business Logic ==========
    ws5 = wb.create_sheet("Business Logic")
    ws5['A1'] = "Business Logic & Model Architecture"
    ws5['A1'].font = TITLE_FONT

    if logic:
        ws5['A3'] = "Detected Model Type"
        ws5['B3'] = logic.get("detected_model_type", "unknown")
        ws5['A5'] = "Cross-Sheet References"
        ws5['B5'] = logic.get("data_flow", {}).get("cross_sheet_references", 0)

        ws5['A8'] = "Financial Patterns Detected"
        for i, p in enumerate(logic.get("financial_patterns", []), 9):
            ws5.cell(row=i, column=1, value=p)

    # ========== Dependencies ==========
    ws6 = wb.create_sheet("Dependencies")
    ws6['A1'] = "Cross-Sheet Dependencies"
    ws6['A1'].font = TITLE_FONT

    if logic and logic.get("data_flow", {}).get("sheet_dependencies"):
        ws6['A3'] = "Top Dependent Sheets"
        for i, (sheet, count) in enumerate(logic["data_flow"]["sheet_dependencies"].items(), 4):
            ws6.cell(row=i, column=1, value=sheet)
            ws6.cell(row=i, column=2, value=count)

    # ========== Named Ranges ==========
    ws7 = wb.create_sheet("Named Ranges")
    ws7['A1'] = "Named Ranges & Variables"
    ws7['A1'].font = TITLE_FONT
    ws7['A3'] = "Name"
    ws7['B3'] = "Hidden"
    ws7['C3'] = "Reference"
    style_header_row(ws7, 3, 1, 3)

    if inspect and inspect.get("named_ranges"):
        for i, nr in enumerate(inspect["named_ranges"], 4):
            ws7.cell(row=i, column=1, value=nr.get("name"))
            ws7.cell(row=i, column=2, value=nr.get("hidden", False))
            ws7.cell(row=i, column=3, value=nr.get("ref", ""))

    # ========== Recommendations ==========
    ws8 = wb.create_sheet("Recommendations")
    ws8['A1'] = "Key Recommendations & Next Steps"
    ws8['A1'].font = TITLE_FONT
    ws8['A3'] = "1. Review all High severity issues first."
    ws8['A4'] = "2. Consider centralizing assumptions if many are scattered."
    ws8['A5'] = "3. For verified values, run xlsx skill's recalc.py on the source file before re-analysis."
    ws8['A6'] = "4. After understanding the model, use the xlsx skill to refactor or recreate it cleanly."

    # Save
    wb.save(output_file)
    print(f"Professional report generated: {output_file}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_report.py <original.xlsx> <report.xlsx> [inspect.json] [audit.json] [logic.json]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    inspect = audit = logic = None

    for i, arg in enumerate(sys.argv[3:], 3):
        try:
            with open(arg) as f:
                data = json.load(f)
                if "sheets" in data:
                    inspect = data
                elif "formula_summary" in data:
                    audit = data
                elif "detected_model_type" in data:
                    logic = data
        except:
            pass

    create_report(input_file, output_file, inspect, audit, logic)

if __name__ == "__main__":
    main()
