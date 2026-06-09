#!/usr/bin/env python3
"""
xlsx-reverse-engineering: Formula Extraction & Basic Dependency Helper
Extracts ALL formulas from the workbook with precise locations.
Performs lightweight dependency parsing (cell refs + cross-sheet).
Outputs structured data + CSV/JSON side artifacts.
Intended to feed the logic archeology phase. Use after inventory and (ideally) after recalc.
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Simple regex for cell references (A1, $A$1, Sheet1!A1, 'Sheet Name'!B2, etc.)
# Not perfect (doesn't handle all edge cases like R1C1 or structured refs) but very effective for most real files.
CELL_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_]+))?!?([$]?[A-Z]{1,3}[$]?[0-9]{1,7})",
    re.IGNORECASE
)

def extract_cell_refs(formula: str):
    """Return list of (sheet, cell) tuples found in the formula string."""
    refs = []
    for match in CELL_REF_RE.finditer(formula):
        sheet = match.group(1) or match.group(2) or None
        cell = match.group(3)
        refs.append((sheet, cell.upper()))
    return refs

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_formulas.py <path-to-xlsx> [max_per_sheet]")
        print("max_per_sheet default=10000 (practically all). Use smaller for very large files.")
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    max_per_sheet = int(sys.argv[2]) if len(sys.argv) > 2 else 10000

    if not input_path.exists():
        print(json.dumps({"error": f"File not found: {input_path}"}))
        sys.exit(1)

    try:
        wb = load_workbook(input_path, data_only=False)
    except Exception as e:
        print(json.dumps({"error": f"openpyxl load failed: {str(e)}"}))
        sys.exit(1)

    all_formulas = []
    formulas_by_sheet = defaultdict(list)
    dep_index = defaultdict(list)  # "Sheet!A1" -> list of formulas that reference it

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    location = f"{sheet_name}!{cell.coordinate}"
                    refs = extract_cell_refs(formula)
                    
                    entry = {
                        "location": location,
                        "formula": formula,
                        "refs": [{"sheet": r[0], "cell": r[1]} for r in refs],
                        "length": len(formula)
                    }
                    all_formulas.append(entry)
                    formulas_by_sheet[sheet_name].append(entry)
                    
                    # Build reverse index (who references this cell)
                    for r in refs:
                        target_sheet = r[0] or sheet_name
                        target = f"{target_sheet}!{r[1]}"
                        dep_index[target].append(location)
                    
                    count += 1
                    if count >= max_per_sheet:
                        break
            if count >= max_per_sheet:
                break

    wb.close()

    # Build summary
    total = len(all_formulas)
    by_length = sorted(all_formulas, key=lambda x: -x["length"])[:5]  # longest for complexity flag

    output = {
        "file": str(input_path),
        "extracted_at": datetime.now().isoformat(),
        "total_formulas": total,
        "sheets_with_formulas": list(formulas_by_sheet.keys()),
        "formulas": all_formulas[:5000] if total > 5000 else all_formulas,  # cap for context safety
        "longest_formulas": [
            {"location": f["location"], "length": f["length"], "preview": f["formula"][:150]} 
            for f in by_length
        ],
        "notes": "Full list may be truncated in this JSON for token safety. See side artifacts for complete data. Refs are heuristically extracted."
    }

    print(json.dumps(output, indent=2, ensure_ascii=False))

    # Side artifacts (very useful for RE)
    out_dir = Path("formulas")
    out_dir.mkdir(exist_ok=True)

    # All formulas as CSV (location,formula)
    with open(out_dir / "all_formulas.csv", "w", encoding="utf-8") as f:
        f.write("location,formula_length,formula\n")
        for entry in all_formulas:
            safe_formula = entry["formula"].replace('"', '""')
            f.write(f'"{entry["location"]}",{entry["length"]},"{safe_formula}"\n')

    # Reverse dependency index (who points to me)
    with open(out_dir / "dependency_index.csv", "w", encoding="utf-8") as f:
        f.write("target,referenced_by_count,example_referencers\n")
        for target, referencers in sorted(dep_index.items()):
            examples = "; ".join(referencers[:3])
            f.write(f'"{target}",{len(referencers)},"{examples}"\n')

    # Per-sheet summary
    with open(out_dir / "formulas_by_sheet.json", "w", encoding="utf-8") as f:
        json.dump({k: [e["location"] for e in v] for k, v in formulas_by_sheet.items()}, f, indent=2)

    print(f"\n[INFO] Complete artifacts in ./formulas/ (all_formulas.csv is the gold source for full list)", file=sys.stderr)
    print(f"[INFO] Total formulas extracted: {total}", file=sys.stderr)
    if total > 5000:
        print("[WARN] JSON output capped at 5000 for safety — use the CSV for everything.", file=sys.stderr)

if __name__ == "__main__":
    main()
