#!/usr/bin/env python3
"""
xlsx-reverse-engineering: Anomaly & Security Pattern Scanner (Full)
Scans for suspicious, obfuscated, or security-relevant patterns based on references/common-spreadsheet-re-patterns.md
Outputs structured anomalies report + Markdown snippet.
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import zipfile

SUSPICIOUS_KEYWORDS = [
    "EVALUATE", "SHELL", "CreateObject", "WEBSERVICE", "INDIRECT", "OFFSET", "CALL",
    "EXEC", "WScript", "powershell", "cmd.exe", "base64", "CHAR(", "CODE("
]

HIDDEN_PATTERNS = ["veryHidden", "state=\"hidden\""]

def scan_xml_for_patterns(unpacked_dir: Path):
    findings = []
    if not unpacked_dir.exists():
        return findings

    for xml_file in unpacked_dir.rglob("*.xml"):
        try:
            content = xml_file.read_text(encoding="utf-8", errors="ignore")
            for kw in SUSPICIOUS_KEYWORDS:
                if kw.lower() in content.lower():
                    findings.append({
                        "type": "xml_suspicious_keyword",
                        "file": str(xml_file.relative_to(unpacked_dir)),
                        "keyword": kw,
                        "snippet": content[max(0, content.lower().find(kw.lower())-30):content.lower().find(kw.lower())+80]
                    })
            for pat in HIDDEN_PATTERNS:
                if pat in content:
                    findings.append({
                        "type": "xml_hidden_state",
                        "file": str(xml_file.relative_to(unpacked_dir)),
                        "pattern": pat
                    })
        except Exception:
            pass
    return findings

def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_anomalies.py <path-to-xlsx> [unpacked_dir]")
        print("If unpacked_dir not provided, it will try ./unpacked-xxx first.")
        sys.exit(1)

    input_path = Path(sys.argv[1]).expanduser().resolve()
    if not input_path.exists():
        print(json.dumps({"error": f"File not found: {input_path}"}))
        sys.exit(1)

    unpacked_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path.cwd() / f"unpacked-{input_path.stem}"

    anomalies = {
        "file": str(input_path),
        "scanned_at": datetime.now().isoformat(),
        "anomalies": [],
        "summary": {}
    }

    # 1. Openpyxl level checks
    try:
        wb = load_workbook(input_path, data_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            state = getattr(ws, 'sheet_state', 'visible')
            if state in ('hidden', 'veryHidden'):
                anomalies["anomalies"].append({
                    "type": "hidden_sheet",
                    "severity": "high" if state == "veryHidden" else "medium",
                    "location": f"Sheet: {sheet_name}",
                    "detail": f"Sheet state = {state}",
                    "recommendation": "Inspect contents via unpack or openpyxl with sheet_state check"
                })

            # Scan for suspicious formulas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value.upper()
                        for kw in SUSPICIOUS_KEYWORDS:
                            if kw in formula:
                                anomalies["anomalies"].append({
                                    "type": "suspicious_formula",
                                    "severity": "high" if kw in ["EVALUATE", "SHELL", "CreateObject"] else "medium",
                                    "location": f"{sheet_name}!{cell.coordinate}",
                                    "formula": cell.value[:150],
                                    "keyword": kw
                                })
        wb.close()
    except Exception as e:
        anomalies["anomalies"].append({"type": "openpyxl_error", "detail": str(e)})

    # 2. Raw XML / unpacked scan
    xml_findings = scan_xml_for_patterns(unpacked_dir)
    for f in xml_findings:
        anomalies["anomalies"].append({
            "type": f["type"],
            "severity": "high" if "veryHidden" in str(f) or f.get("keyword") in ["EVALUATE", "SHELL"] else "medium",
            **f
        })

    # Summary
    high = sum(1 for a in anomalies["anomalies"] if a.get("severity") == "high")
    medium = sum(1 for a in anomalies["anomalies"] if a.get("severity") == "medium")
    anomalies["summary"] = {
        "total_anomalies": len(anomalies["anomalies"]),
        "high_severity": high,
        "medium_severity": medium,
        "recommendation": "Review high severity items first. Cross-reference with REVERSE-EXCEL.md"
    }

    print(json.dumps(anomalies, indent=2, ensure_ascii=False))

    # Write artifacts
    out_dir = Path("anomalies")
    out_dir.mkdir(exist_ok=True)
    with open(out_dir / "anomalies.json", "w", encoding="utf-8") as f:
        json.dump(anomalies, f, indent=2, ensure_ascii=False)

    with open(out_dir / "anomalies_report.md", "w", encoding="utf-8") as f:
        f.write(f"# Anomaly Scan Report\n\n**File**: {input_path}\n**Scanned**: {anomalies['scanned_at']}\n\n")
        f.write(f"**Summary**: {anomalies['summary']['total_anomalies']} anomalies ({high} high, {medium} medium)\n\n")
        for a in anomalies["anomalies"]:
            f.write(f"### {a.get('type', 'unknown').upper()} - {a.get('severity', 'medium').upper()}\n")
            f.write(f"**Location**: {a.get('location', a.get('file', 'N/A'))}\n")
            if 'formula' in a:
                f.write(f"**Formula**: `{a['formula']}`\n")
            if 'keyword' in a:
                f.write(f"**Matched**: {a['keyword']}\n")
            f.write(f"**Recommendation**: {a.get('recommendation', 'Investigate manually')}\n\n")

    print(f"\n[INFO] Full report in ./anomalies/anomalies_report.md and anomalies.json", file=sys.stderr)

if __name__ == "__main__":
    main()
