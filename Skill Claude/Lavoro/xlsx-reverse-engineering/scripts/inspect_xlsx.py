#!/usr/bin/env python3
"""
inspect_xlsx.py — Structural Analysis (Phase 1)
Supports --data-only to load calculated values (for files that have been recalculated).
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def main():
    parser = argparse.ArgumentParser(description="Structural inspection of Excel file")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--data-only", action="store_true", help="Load with data_only=True for calculated values (file must have been recalculated)")
    parser.add_argument("--pretty", action="store_true", help="Pretty print JSON")
    args = parser.parse_args()

    input_path = Path(args.file).expanduser().resolve()
    if not input_path.exists():
        print(json.dumps({"error": "File not found"}))
        sys.exit(1)

    wb = load_workbook(input_path, data_only=False)
    wb_values = None
    if args.data_only:
        try:
            wb_values = load_workbook(input_path, data_only=True)
        except Exception as e:
            print(f"Warning: Could not load with data_only=True: {e}", file=sys.stderr)

    result = {
        "file": str(input_path),
        "data_only_mode": args.data_only,
        "scanned_at": datetime.now().isoformat(),
        "summary": {
            "total_sheets": len(wb.sheetnames),
            "total_formulas": 0,
            "formula_density": 0,
            "has_macros": input_path.suffix.lower() == ".xlsm",
            "has_tables": False,
            "hidden_sheets": 0,
            "very_hidden_sheets": 0,
        },
        "sheets": [],
        "named_ranges": [],
    }

    total_cells = 0
    total_formulas = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_v = wb_values[sheet_name] if wb_values else None
        state = getattr(ws, "sheet_state", "visible")

        if state == "hidden":
            result["summary"]["hidden_sheets"] += 1
        elif state == "veryHidden":
            result["summary"]["very_hidden_sheets"] += 1

        formula_count = 0
        samples = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1000), max_col=min(ws.max_column, 50)):
            for cell in row:
                total_cells += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    total_formulas += 1
                    if len(samples) < 3:
                        val = None
                        if ws_v:
                            try:
                                val = ws_v[cell.coordinate].value
                            except:
                                pass
                        samples.append({
                            "cell": cell.coordinate,
                            "formula": cell.value[:100],
                            "value": str(val)[:50] if val is not None else None
                        })

        has_tables = len(getattr(ws, "tables", {})) > 0
        if has_tables:
            result["summary"]["has_tables"] = True

        result["sheets"].append({
            "name": sheet_name,
            "state": state,
            "dimensions": f"{get_column_letter(1)}1:{get_column_letter(ws.max_column)}{ws.max_row}",
            "formula_count": formula_count,
            "formula_samples": samples,
            "has_tables": has_tables,
            "protected": bool(getattr(ws, "protection", None) and ws.protection.sheet),
        })

    result["summary"]["total_formulas"] = total_formulas
    if total_cells > 0:
        result["summary"]["formula_density"] = round(total_formulas / total_cells * 100, 2)

    for name_str in wb.defined_names:
        dn = wb.defined_names[name_str]
        result["named_ranges"].append({
            "name": getattr(dn, "name", name_str),
            "hidden": bool(getattr(dn, "hidden", False)),
            "ref": getattr(dn, "attr_text", None) or str(getattr(dn, "value", ""))
        })

    wb.close()
    if wb_values:
        wb_values.close()

    if args.pretty:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(json.dumps(result))

if __name__ == "__main__":
    main()
