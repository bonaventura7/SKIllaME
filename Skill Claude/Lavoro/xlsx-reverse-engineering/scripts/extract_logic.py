#!/usr/bin/env python3
"""
extract_logic.py — Business Logic Extraction (Phase 3)
Supports --data-only for value-based assumption detection.
"""

import json
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    parser.add_argument("--data-only", action="store_true")
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()

    input_path = Path(args.file).expanduser().resolve()
    wb = load_workbook(input_path, data_only=False)
    wb_values = load_workbook(input_path, data_only=True) if args.data_only else None

    all_formulas = ""
    cross_sheet_refs = 0
    sheet_deps = defaultdict(int)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    all_formulas += " " + cell.value
                    if "!" in cell.value:
                        cross_sheet_refs += 1
                        for m in re.findall(r"'?([^'!]+)'?!", cell.value):
                            sheet_deps[m] += 1

    result = {
        "file": str(input_path),
        "data_only_mode": args.data_only,
        "scanned_at": datetime.now().isoformat(),
        "detected_model_type": "unknown",
        "financial_patterns": [],
        "assumptions": [],
        "data_flow": {
            "cross_sheet_references": cross_sheet_refs,
            "sheet_dependencies": dict(sorted(sheet_deps.items(), key=lambda x: -x[1])[:8])
        },
        "risk_indicators": []
    }

    # Model type detection (same as before, kept for brevity)
    text = " ".join(wb.sheetnames).lower() + " " + all_formulas.lower()
    if any(k in text for k in ["dcf", "wacc", "terminal"]):
        result["detected_model_type"] = "dcf_valuation"
    elif any(k in text for k in ["budget", "forecast"]):
        result["detected_model_type"] = "budget_forecast"
    elif any(k in ["revenue", "cogs", "ebitda"] for k in text.split()):
        result["detected_model_type"] = "three_statement_financial_model"

    # Assumptions using values if data_only
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_v = wb_values[sheet_name] if wb_values else None
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=8):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    val = cell.value
                    if ws_v:
                        try:
                            val = ws_v[cell.coordinate].value or cell.value
                        except:
                            pass
                    if 0 < abs(val) < 1.1 or (0 < val < 200 and val % 1 != 0):
                        result["assumptions"].append({
                            "location": f"{sheet_name}!{cell.coordinate}",
                            "value": val
                        })

    if cross_sheet_refs > 60:
        result["risk_indicators"].append("High cross-sheet dependency count")

    wb.close()
    if wb_values:
        wb_values.close()

    if args.pretty:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(json.dumps(result))

if __name__ == "__main__":
    main()
